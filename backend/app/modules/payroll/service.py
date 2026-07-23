"""Payroll Management — service layer (business logic & orchestration).

Implements the business logic of the Payroll Management API Contract.
All database access is performed strictly via repositories and session queries.
"""

from __future__ import annotations

import io
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, time, timedelta
from decimal import Decimal
from typing import TYPE_CHECKING

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions.base import ConflictException, NotFoundException, ValidationException
from app.modules.attendance.models import AttendanceDay
from app.modules.audit.constants import ActionType
from app.modules.audit.service import AuditService
from app.modules.employee.models.employee import Employee
from app.modules.employee.repository import EmployeeRepository
from app.modules.leave.models.leave import LeaveRequest
from app.modules.notifications.constants import NotificationType
from app.modules.payroll.constants import AttendanceMode, PaymentStatus, WorkingHourType
from app.modules.payroll.exceptions import (
    AdjustmentExistsException,
    AdjustmentNotFoundException,
    ComputedRowNotFoundException,
    CycleExistsException,
    CycleFinalizedException,
    CycleNotFoundException,
    EmployeeNotFoundException,
    FinalizedRunNotFoundException,
    PayrollAlreadyFinalizedException,
    PayrollGroupInUseException,
    PayrollGroupNameExistsException,
    PayrollGroupNotFoundException,
    PayrollNotFinalizedException,
)
from app.modules.payroll.models import (
    AttendanceAdjustment,
    AttendanceAdjustmentExtraHours,
    AttendanceAdjustmentPenalty,
    EmployeePayrollGroupAssignment,
    FinalizedPayrollRun,
    PayrollColumnSetting,
    PayrollComputedRow,
    PayrollFinalization,
    PayrollFinalizationEmployee,
    PayrollGroup,
    PayrollSalaryCycle,
    PayrollSetting,
)
from app.modules.payroll.repository import (
    AttendanceAdjustmentExtraHoursRepository,
    AttendanceAdjustmentPenaltyRepository,
    AttendanceAdjustmentRepository,
    EmployeePayrollGroupAssignmentRepository,
    FinalizedPayrollRunRepository,
    PayrollColumnSettingRepository,
    PayrollComputedRowRepository,
    PayrollFinalizationRepository,
    PayrollGroupRepository,
    PayrollSalaryCycleRepository,
    PayrollSettingRepository,
)
from app.modules.payroll.schemas import (
    AttendanceAdjustmentCreateSchema,
    AttendanceAdjustmentExtraHoursCreateSchema,
    AttendanceAdjustmentPenaltyCreateSchema,
    AttendanceAdjustmentUpdateSchema,
    BulkAttendanceAdjustmentBatchUpdateResponseSchema,
    BulkAttendanceAdjustmentBatchUpdateSchema,
    BulkAttendanceAdjustmentMatrixItemSchema,
    BulkAttendanceAdjustmentMatrixResponseSchema,
    PayrollFinalizationCancelSchema,
    PayrollFinalizationCreateSchema,
    PayrollFinalizationEmployeeSchema,
    PayrollFinalizationListResponse,
    PayrollFinalizationPaySchema,
    PayrollFinalizationResponseSchema,
    BulkAttendanceAdjustmentResetSchema,
    EmployeeGroupAssignRequestSchema,
    GroupEmployeesResponseSchema,
    PayrollColumnSettingsReplaceSchema,
    PayrollCycleCreateSchema,
    PayrollCycleUpdateSchema,
    PayrollGroupAssignEmployeesSchema,
    PayrollGroupCreateSchema,
    PayrollGroupResponseSchema,
    PayrollGroupUpdateSchema,
    PayrollProcessItemResultSchema,
    PayrollProcessRequestSchema,
    PayrollProcessResponseSchema,
    PayrollSettingUpdateSchema,
    PayrollSummaryResponseSchema,
    PayslipResponseSchema,
    PayslipSectionItemSchema,
    RecordPaymentRequestSchema,
)
from app.shared.schemas.pagination import PaginationMeta
from app.modules.rbac.repository import UserRepository
from app.modules.settlements.models import (
    ArrearsTransaction,
    EmployeeArrears,
    EmployeeLoanAdvance,
    LoanAdvanceTransaction,
)
from app.shared.base.service import BaseService
from app.shared.schemas.pagination import PaginatedResponse
from app.shared.utils.datetime import utcnow

if TYPE_CHECKING:
    from app.modules.notifications.service import NotificationService


# ``AttendanceAdjustmentRepository.search`` is called with ``page_size=100``; the bulk
# prefetch reproduces that per-employee page bound exactly.
_ADJUSTMENT_PAGE_SIZE = 100


@dataclass(frozen=True)
class EmployeePayrollInputs:
    """Every already-fetched input the calculator needs for ONE employee.

    Produced by :meth:`PayrollService._prefetch_batch_inputs`. Holding it lets
    :meth:`PayrollService._calculate_employee_payroll` stay pure — it performs no
    database access of its own.
    """

    assignment: EmployeePayrollGroupAssignment | None = None
    attendance_days: list[AttendanceDay] = field(default_factory=list)
    adjustments: list[AttendanceAdjustment] = field(default_factory=list)
    leaves: list[LeaveRequest] = field(default_factory=list)
    penalties: list[AttendanceAdjustmentPenalty] = field(default_factory=list)
    extra_hours: list[AttendanceAdjustmentExtraHours] = field(default_factory=list)
    loans: list[EmployeeLoanAdvance] = field(default_factory=list)
    arrears: EmployeeArrears | None = None


@dataclass(frozen=True)
class PayrollBatchInputs:
    """Bulk-loaded payroll inputs for a whole employee set, keyed by ``employee_id``."""

    assignments: dict[int, EmployeePayrollGroupAssignment] = field(default_factory=dict)
    attendance_days: dict[int, list[AttendanceDay]] = field(default_factory=dict)
    adjustments: dict[int, list[AttendanceAdjustment]] = field(default_factory=dict)
    leaves: dict[int, list[LeaveRequest]] = field(default_factory=dict)
    penalties: dict[int, list[AttendanceAdjustmentPenalty]] = field(default_factory=dict)
    extra_hours: dict[int, list[AttendanceAdjustmentExtraHours]] = field(default_factory=dict)
    loans: dict[int, list[EmployeeLoanAdvance]] = field(default_factory=dict)
    arrears: dict[int, EmployeeArrears] = field(default_factory=dict)

    def for_employee(self, employee_id: int) -> EmployeePayrollInputs:
        """Slice out the inputs belonging to a single employee."""
        return EmployeePayrollInputs(
            assignment=self.assignments.get(employee_id),
            attendance_days=self.attendance_days.get(employee_id, []),
            adjustments=self.adjustments.get(employee_id, []),
            leaves=self.leaves.get(employee_id, []),
            penalties=self.penalties.get(employee_id, []),
            extra_hours=self.extra_hours.get(employee_id, []),
            loans=self.loans.get(employee_id, []),
            arrears=self.arrears.get(employee_id),
        )


class PayrollService(BaseService):
    """Business rules engine and orchestration service for Payroll Management."""

    def __init__(self, session: AsyncSession) -> None:
        super().__init__(session)
        # PayrollRepositories
        self.settings = PayrollSettingRepository(session)
        self.groups = PayrollGroupRepository(session)
        self.assignments = EmployeePayrollGroupAssignmentRepository(session)
        self.cycles = PayrollSalaryCycleRepository(session)
        self.columns = PayrollColumnSettingRepository(session)
        self.runs = FinalizedPayrollRunRepository(session)
        self.finalizations = PayrollFinalizationRepository(session)
        self.computed_rows = PayrollComputedRowRepository(session)
        self.adjustments = AttendanceAdjustmentRepository(session)
        self.penalties = AttendanceAdjustmentPenaltyRepository(session)
        self.extra_hours = AttendanceAdjustmentExtraHoursRepository(session)

        # Cross-module repositories & services
        self.employees = EmployeeRepository(session)
        self.users = UserRepository(session)
        self.audit = AuditService(session)

        # Cross-module notifier (constructed lazily — see _get_notifier).
        self.notifications: NotificationService | None = None

    # --- Helper & Guard Methods ----------------------------------------------

    def _get_notifier(self) -> NotificationService:
        """Return the notifications service, importing lazily to avoid module-level coupling."""
        if self.notifications is None:
            from app.modules.notifications.service import NotificationService

            self.notifications = NotificationService(self.session)
        return self.notifications

    async def _validate_employee(self, org_id: int, employee_id: int) -> Employee:
        """Validate employee existence and active status in organization context."""
        employee = await self.employees.get_by_id(employee_id)
        if not employee or employee.org_id != org_id or employee.is_deleted:
            raise EmployeeNotFoundException()
        return employee

    async def _validate_payroll_group(self, org_id: int, group_id: int) -> PayrollGroup:
        """Validate payroll group existence and status in organization context."""
        group = await self.groups.get_by_id_in_org(org_id, group_id)
        if not group:
            raise PayrollGroupNotFoundException()
        return group

    async def _get_or_create_settings(self, org_id: int) -> PayrollSetting:
        """Get or create default payroll settings for the organization."""
        settings = await self.settings.get_by_org(org_id)
        if not settings:
            settings = await self.settings.create(
                {
                    "org_id": org_id,
                    "working_hour_type": WorkingHourType.FIXED.value,
                    "full_day_working_hours": time(9, 0),
                    "half_day_working_hours": time(4, 30),
                    "attendance_mode": AttendanceMode.CONSIDER_ALL_PUNCH.value,
                    "off_day_compensation": "paid",
                    "off_day_wage_multiplier": Decimal("1.0"),
                    "daily_wage_formula": "calendar_days",
                    "overtime_type": "multiplier",
                    "overtime_hourly_multiplier": Decimal("1.5"),
                    "overtime_buffer_period": time(0, 30),
                    "overtime_period_interval": "daily",
                    "full_day_penalty_enabled": False,
                    "half_day_penalty_enabled": False,
                    "late_coming_penalty_enabled": False,
                    "grace_time": time(0, 15),
                }
            )
        return settings

    # --- 1. Payroll Configuration --------------------------------------------

    async def get_settings(self, org_id: int) -> PayrollSetting:
        """Retrieve organization-wide calculation settings."""
        return await self._get_or_create_settings(org_id)

    async def update_settings(
        self, org_id: int, payload: PayrollSettingUpdateSchema, user_id: int
    ) -> PayrollSetting:
        """Update organization-wide calculation settings."""
        settings = await self._get_or_create_settings(org_id)
        update_data = payload.model_dump(exclude_unset=True)

        async with self.transaction():
            updated = await self.settings.update(settings, update_data)
            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="settings",
                action_type=ActionType.UPDATE,
                title="Update Payroll Settings",
                description="Updated organization-level payroll settings parameters.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return updated

    # --- 2. Payroll Groups (Salary Structures) ------------------------------

    # --- 2. Payroll Groups (Salary Structures) ------------------------------

    async def create_group(
        self, org_id: int, payload: PayrollGroupCreateSchema, user_id: int
    ) -> PayrollGroupResponseSchema:
        """Create a new payroll group and enforce default policy constraints per payroll type."""
        async with self.transaction():
            # Validate name uniqueness among non-deleted groups in this org
            name_exists = await self.groups.name_exists(org_id, payload.name)
            if name_exists:
                raise PayrollGroupNameExistsException()

            if payload.is_default:
                await self.groups.clear_defaults_for_type(org_id, payload.payroll_type.value)

            group = await self.groups.create(
                {
                    "org_id": org_id,
                    "name": payload.name,
                    "payroll_type": payload.payroll_type.value,
                    "is_default": payload.is_default,
                    "is_deleted": False,
                    "created_by": user_id,
                    "updated_by": user_id,
                }
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="groups",
                action_type=ActionType.INSERT,
                title="Create Payroll Group",
                description=f"Created payroll group '{payload.name}'.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        resp = PayrollGroupResponseSchema.model_validate(group)
        resp.employee_count = 0
        return resp

    async def list_groups(
        self,
        org_id: int,
        *,
        search: str | None = None,
        payroll_type: str | None = None,
        is_default: bool | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
        page: int = 1,
        page_size: int = 25,
    ) -> PaginatedResponse[PayrollGroupResponseSchema]:
        """List paginated payroll groups with search, filters, sorting, and employee count."""
        groups = await self.groups.search(
            org_id,
            search=search,
            payroll_type=payroll_type,
            is_default=is_default,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            page_size=page_size,
        )
        total = await self.groups.search_count(
            org_id, search=search, payroll_type=payroll_type, is_default=is_default
        )

        group_ids = [g.id for g in groups]
        emp_counts = await self.assignments.count_by_groups(group_ids)

        items = []
        for g in groups:
            item = PayrollGroupResponseSchema.model_validate(g)
            item.employee_count = emp_counts.get(g.id, 0)
            items.append(item)

        return self.paginate(items, page=page, page_size=page_size, total_records=total)

    async def get_group(self, org_id: int, group_id: int) -> PayrollGroupResponseSchema:
        """Get payroll group details scoped by org."""
        group = await self._validate_payroll_group(org_id, group_id)
        emp_count = await self.assignments.count_by_group(group_id)
        resp = PayrollGroupResponseSchema.model_validate(group)
        resp.employee_count = emp_count
        return resp

    async def update_group(
        self, org_id: int, group_id: int, payload: PayrollGroupUpdateSchema, user_id: int
    ) -> PayrollGroupResponseSchema:
        """Update payroll group and enforce one default per payroll type."""
        group = await self._validate_payroll_group(org_id, group_id)
        update_data = payload.model_dump(exclude_unset=True)

        async with self.transaction():
            if "name" in update_data and update_data["name"] != group.name:
                name_exists = await self.groups.name_exists(
                    org_id, update_data["name"], exclude_id=group_id
                )
                if name_exists:
                    raise PayrollGroupNameExistsException()

            target_type = update_data.get("payroll_type", group.payroll_type)
            if isinstance(target_type, Enum):
                target_type = target_type.value

            if update_data.get("is_default"):
                await self.groups.clear_defaults_for_type(org_id, target_type, exclude_id=group_id)

            update_data["updated_by"] = user_id
            if "payroll_type" in update_data and isinstance(update_data["payroll_type"], Enum):
                update_data["payroll_type"] = update_data["payroll_type"].value

            updated = await self.groups.update(group, update_data)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="groups",
                action_type=ActionType.UPDATE,
                title="Update Payroll Group",
                description=f"Updated payroll group ID {group_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        emp_count = await self.assignments.count_by_group(group_id)
        resp = PayrollGroupResponseSchema.model_validate(updated)
        resp.employee_count = emp_count
        return resp

    async def delete_group(self, org_id: int, group_id: int, user_id: int) -> None:
        """Soft delete a payroll group enforcing Business Rules."""
        group = await self._validate_payroll_group(org_id, group_id)

        # Rule: Cannot delete Default Group
        if group.is_default:
            raise PayrollGroupInUseException("Default payroll group cannot be deleted.")

        # Rule: Cannot delete Group with employees unless reassigned
        emp_count = await self.assignments.count_by_group(group_id)
        if emp_count > 0:
            raise PayrollGroupInUseException(
                "Cannot delete payroll group with assigned employees. Reassign employees before deletion."
            )

        # Check if assigned in Employee table
        stmt_emp = select(Employee.employee_id).where(
            Employee.payroll_group_id == group_id, Employee.is_deleted.is_(False)
        )
        has_employees = (await self.session.execute(stmt_emp.limit(1))).first() is not None
        if has_employees:
            raise PayrollGroupInUseException(
                "Cannot delete payroll group with assigned employees. Reassign employees before deletion."
            )

        # Check if referenced by cycles
        stmt_cycle = select(PayrollSalaryCycle.id).where(
            PayrollSalaryCycle.payroll_group_id == group_id
        )
        has_cycles = (await self.session.execute(stmt_cycle.limit(1))).first() is not None
        if has_cycles:
            raise PayrollGroupInUseException("Group has active salary cycles.")

        # Check if referenced by runs
        stmt_run = select(FinalizedPayrollRun.id).where(
            FinalizedPayrollRun.payroll_group_id == group_id,
            FinalizedPayrollRun.is_definalized.is_(False),
        )
        has_runs = (await self.session.execute(stmt_run.limit(1))).first() is not None
        if has_runs:
            raise PayrollGroupInUseException("Group has active finalized runs.")

        async with self.transaction():
            await self.groups.update(group, {"is_deleted": True, "updated_by": user_id})
            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="groups",
                action_type=ActionType.DELETE,
                title="Delete Payroll Group",
                description=f"Soft deleted payroll group ID {group_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

    async def assign_employees_to_group(
        self,
        org_id: int,
        group_id: int,
        payload: PayrollGroupAssignEmployeesSchema,
        user_id: int,
    ) -> int:
        """Batch assign employees to a specific payroll group."""
        group = await self._validate_payroll_group(org_id, group_id)
        assigned_count = 0

        async with self.transaction():
            for emp_id in payload.employee_ids:
                existing = await self.assignments.get_by_employee(emp_id)
                prev_gid = existing.payroll_group_id if existing else None

                if existing:
                    await self.assignments.update(
                        existing,
                        {
                            "payroll_group_id": group_id,
                            "salary_type": payload.salary_type.value,
                            "previous_group_id": prev_gid,
                            "assigned_by": user_id,
                            "assigned_at": utcnow(),
                        },
                    )
                else:
                    await self.assignments.create(
                        {
                            "employee_id": emp_id,
                            "payroll_group_id": group_id,
                            "salary_type": payload.salary_type.value,
                            "previous_group_id": None,
                            "assigned_by": user_id,
                            "assigned_at": utcnow(),
                        }
                    )

                # Update Employee table FK
                emp = await self.employees.get_by_id(emp_id)
                if emp and emp.org_id == org_id:
                    await self.employees.update(emp, {"payroll_group_id": group_id})

                assigned_count += 1

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="assignments",
                action_type=ActionType.UPDATE,
                title="Assign Employees to Group",
                description=f"Assigned {assigned_count} employees to payroll group '{group.name}'.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        return assigned_count

    async def get_group_employees(
        self,
        org_id: int,
        group_id: int,
        page: int = 1,
        page_size: int = 25,
        search: str | None = None,
    ) -> GroupEmployeesResponseSchema:
        """Get paginated list of employees assigned to group_id."""
        group = await self._validate_payroll_group(org_id, group_id)
        total, items = await self.assignments.get_assigned_employees(
            org_id=org_id, group_id=group_id, page=page, page_size=page_size, search=search
        )
        return GroupEmployeesResponseSchema(
            payroll_group_id=group.id,
            payroll_group_name=group.name,
            total_employees=total,
            items=items,
        )

    # --- 3. Employee Group Assignment ---------------------------------------

    async def assign_group(
        self, org_id: int, employee_id: int, payload: EmployeeGroupAssignRequestSchema, user_id: int
    ) -> EmployeePayrollGroupAssignment:
        """Assign employee to payroll group, recording assignment history."""
        employee = await self._validate_employee(org_id, employee_id)
        await self._validate_payroll_group(org_id, payload.payroll_group_id)

        async with self.transaction():
            existing = await self.assignments.get_by_employee(employee_id)
            previous_group_id = existing.payroll_group_id if existing else None

            if existing:
                assignment = await self.assignments.update(
                    existing,
                    {
                        "payroll_group_id": payload.payroll_group_id,
                        "salary_type": payload.salary_type.value,
                        "previous_group_id": previous_group_id,
                        "assigned_by": user_id,
                        "assigned_at": utcnow(),
                    },
                )
            else:
                assignment = await self.assignments.create(
                    {
                        "employee_id": employee_id,
                        "payroll_group_id": payload.payroll_group_id,
                        "salary_type": payload.salary_type.value,
                        "previous_group_id": None,
                        "assigned_by": user_id,
                        "assigned_at": utcnow(),
                    }
                )

            # Keep Employee master record sync'd
            await self.employees.update(
                employee,
                {
                    "payroll_group_id": payload.payroll_group_id,
                    "salary_type": payload.salary_type.value.capitalize(),
                },
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="assignments",
                action_type=ActionType.ASSIGN,
                title="Assign Payroll Group",
                description=f"Assigned employee ID {employee_id} to group ID "
                    f"{payload.payroll_group_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=employee_id,
            )
        return assignment

    async def get_employee_assignment(
        self, org_id: int, employee_id: int
    ) -> EmployeePayrollGroupAssignment:
        """Get current employee assignment details."""
        await self._validate_employee(org_id, employee_id)
        assignment = await self.assignments.get_by_employee(employee_id)
        if not assignment:
            raise NotFoundException("Employee has no group assignment.")
        return assignment

    # --- 4. Group Column Settings -------------------------------------------

    async def list_columns(self, org_id: int, group_id: int) -> list[PayrollColumnSetting]:
        """List column settings for a payroll group."""
        await self._validate_payroll_group(org_id, group_id)
        return await self.columns.get_by_group(group_id)

    async def replace_columns(
        self, org_id: int, group_id: int, payload: PayrollColumnSettingsReplaceSchema, user_id: int
    ) -> list[PayrollColumnSetting]:
        """Replace column settings layout for a group inside a transaction."""
        await self._validate_payroll_group(org_id, group_id)

        columns_data = [col.model_dump() for col in payload.columns]
        async with self.transaction():
            replaced = await self.columns.replace_columns(group_id, columns_data, user_id)
            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="columns",
                action_type=ActionType.UPDATE,
                title="Replace Column Settings",
                description=f"Replaced column settings layout for group ID {group_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return replaced

    # --- 5. Payroll Cycles ---------------------------------------------------

    async def create_cycle(
        self, org_id: int, payload: PayrollCycleCreateSchema, user_id: int
    ) -> PayrollSalaryCycle:
        """Create a new payroll period cycle."""
        await self._validate_payroll_group(org_id, payload.payroll_group_id)

        async with self.transaction():
            existing = await self.cycles.get_cycle(payload.payroll_group_id, payload.cycle_date)
            if existing:
                raise CycleExistsException()

            cycle = await self.cycles.create(
                {
                    "payroll_group_id": payload.payroll_group_id,
                    "cycle_date": payload.cycle_date,
                    "is_finalized": False,
                    "created_by": user_id,
                }
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="cycles",
                action_type=ActionType.INSERT,
                title="Create Salary Cycle",
                description=f"Created cycle for group ID {payload.payroll_group_id} date "
                    f"{payload.cycle_date}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return cycle

    async def list_cycles(
        self,
        org_id: int,
        group_id: int | None,
        is_finalized: bool | None,
        page: int,
        page_size: int,
    ) -> PaginatedResponse[PayrollSalaryCycle]:
        """List paginated cycles."""
        cycles = await self.cycles.search(
            org_id,
            payroll_group_id=group_id,
            is_finalized=is_finalized,
            page=page,
            page_size=page_size,
        )
        total = await self.cycles.search_count(
            org_id, payroll_group_id=group_id, is_finalized=is_finalized
        )
        return self.paginate(cycles, page=page, page_size=page_size, total_records=total)

    async def update_cycle(
        self, org_id: int, cycle_id: int, payload: PayrollCycleUpdateSchema, user_id: int
    ) -> PayrollSalaryCycle:
        """Update cycle date only if cycle is not finalized."""
        cycle = await self.cycles.get_by_id(cycle_id)
        if not cycle:
            raise CycleNotFoundException()

        # Scoping validation
        await self._validate_payroll_group(org_id, cycle.payroll_group_id)

        if cycle.is_finalized:
            raise CycleFinalizedException()

        async with self.transaction():
            updated = await self.cycles.update(cycle, {"cycle_date": payload.cycle_date})
            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="cycles",
                action_type=ActionType.UPDATE,
                title="Update Salary Cycle",
                description=f"Updated cycle ID {cycle_id} target date to {payload.cycle_date}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return updated

    # --- 6. Payroll Computation Engine ---------------------------------------

    async def _prefetch_batch_inputs(
        self,
        org_id: int,
        employee_ids: list[int],
        cycle_from: date,
        cycle_to: date,
    ) -> PayrollBatchInputs:
        """Bulk-load every calculator input for the whole employee set.

        Issues a fixed number of ``WHERE ... IN (:employee_ids)`` queries (independent of
        the number of employees) and buckets the results by ``employee_id``, so the
        per-employee compute loop can run entirely in memory.
        """
        if not employee_ids:
            return PayrollBatchInputs()

        # 1. Group assignments (``employee_id`` is unique on the table).
        assignment_rows = await self.assignments.get_by_employees(employee_ids)
        assignments = {a.employee_id: a for a in assignment_rows}

        # 2. Attendance days for the cycle.
        attendance_days: dict[int, list[AttendanceDay]] = defaultdict(list)
        stmt_att = select(AttendanceDay).where(
            AttendanceDay.org_id == org_id,
            AttendanceDay.employee_id.in_(employee_ids),
            AttendanceDay.attendance_date >= cycle_from,
            AttendanceDay.attendance_date <= cycle_to,
        )
        for day in (await self.session.execute(stmt_att)).scalars().all():
            attendance_days[day.employee_id].append(day)

        # 3. Manual attendance adjustments. Bucketed newest-first and capped per employee,
        #    reproducing the single-employee ``adjustments.search(..., page_size=100)`` page.
        adjustments: dict[int, list[AttendanceAdjustment]] = defaultdict(list)
        adjustment_rows = await self.adjustments.get_adjustments_for_employees(
            org_id, employee_ids, cycle_from, cycle_to
        )
        for adj in adjustment_rows:
            bucket = adjustments[adj.employee_id]
            if len(bucket) < _ADJUSTMENT_PAGE_SIZE:
                bucket.append(adj)

        # 4. Approved leave requests overlapping the cycle.
        leaves: dict[int, list[LeaveRequest]] = defaultdict(list)
        stmt_leave = select(LeaveRequest).where(
            LeaveRequest.employee_id.in_(employee_ids),
            LeaveRequest.status == "approved",
            LeaveRequest.start_date <= cycle_to,
            LeaveRequest.end_date >= cycle_from,
        )
        for leave in (await self.session.execute(stmt_leave)).scalars().all():
            leaves[leave.employee_id].append(leave)

        # 5. Active manual penalties.
        penalties: dict[int, list[AttendanceAdjustmentPenalty]] = defaultdict(list)
        penalty_rows = await self.penalties.get_penalties_for_employees(
            employee_ids, cycle_from, cycle_to
        )
        for pen in penalty_rows:
            penalties[pen.employee_id].append(pen)

        # 6. Extra hours logs.
        extra_hours: dict[int, list[AttendanceAdjustmentExtraHours]] = defaultdict(list)
        extra_hour_rows = await self.extra_hours.get_extra_hours_for_employees(
            employee_ids, cycle_from, cycle_to
        )
        for record in extra_hour_rows:
            extra_hours[record.employee_id].append(record)

        # 7. Active loans / advances.
        loans: dict[int, list[EmployeeLoanAdvance]] = defaultdict(list)
        stmt_loans = select(EmployeeLoanAdvance).where(
            EmployeeLoanAdvance.org_id == org_id,
            EmployeeLoanAdvance.employee_id.in_(employee_ids),
            EmployeeLoanAdvance.status == "active",
        ).order_by(EmployeeLoanAdvance.id)
        for loan in (await self.session.execute(stmt_loans)).scalars().all():
            loans[loan.employee_id].append(loan)

        # 8. Arrears (unique per org + employee).
        stmt_arr = select(EmployeeArrears).where(
            EmployeeArrears.org_id == org_id,
            EmployeeArrears.employee_id.in_(employee_ids),
        )
        arrears = {
            rec.employee_id: rec for rec in (await self.session.execute(stmt_arr)).scalars().all()
        }

        return PayrollBatchInputs(
            assignments=assignments,
            attendance_days=dict(attendance_days),
            adjustments=dict(adjustments),
            leaves=dict(leaves),
            penalties=dict(penalties),
            extra_hours=dict(extra_hours),
            loans=dict(loans),
            arrears=arrears,
        )

    async def _resolve_employees(
        self, org_id: int, payload: PayrollProcessRequestSchema
    ) -> list[Employee]:
        """Resolve the employee set a processing request targets."""
        stmt_emp = select(Employee).where(
            Employee.org_id == org_id, Employee.is_deleted.is_(False)
        )
        if payload.employee_ids:
            stmt_emp = stmt_emp.where(Employee.employee_id.in_(payload.employee_ids))
        else:
            stmt_emp = stmt_emp.where(Employee.payroll_group_id == payload.payroll_group_id)
        return list((await self.session.execute(stmt_emp)).scalars().all())

    def _calculate_employee_payroll(
        self,
        employee: Employee,
        cycle_from: date,
        cycle_to: date,
        settings: PayrollSetting,
        user_id: int,
        inputs: EmployeePayrollInputs,
    ) -> dict:
        """Pure computation engine for an employee's salary and components.

        Performs **no** database access: every input arrives pre-fetched in ``inputs``.
        """
        # 1. Verify group assignment
        assignment = inputs.assignment
        if not assignment:
            raise ValidationException(f"Employee {employee.employee_id} not assigned to payroll "
                "group.")

        monthly_salary = employee.monthly_salary or Decimal("0.00")
        salary_type = assignment.salary_type

        # 2. Base metrics
        total_days = (cycle_to - cycle_from).days + 1
        if total_days <= 0:
            raise ValidationException("Invalid cycle date range.")

        # 3. Attendance days (pre-fetched)
        att_days = inputs.attendance_days

        # Map counts
        fd_count = 0
        hd_count = 0
        wo_count = 0
        total_working_mins = 0
        overtime_mins = 0
        late_mins = 0

        for day in att_days:
            total_working_mins += day.total_working_minutes
            overtime_mins += day.overtime_minutes
            late_mins += day.late_minutes
            if day.status == "present":
                fd_count += 1
            elif day.status == "half_day":
                hd_count += 1
            elif day.status == "week_off":
                wo_count += 1

        # 4. Overwrite counts via manual adjustments (pre-fetched)
        adjustments = inputs.adjustments
        adj_map = {adj.attendance_date: adj for adj in adjustments}

        # Deduct adjusted dates from baseline before adding back adjusted
        for day in att_days:
            if day.attendance_date in adj_map:
                # remove from baseline
                if day.status == "present":
                    fd_count -= 1
                elif day.status == "half_day":
                    hd_count -= 1
                elif day.status == "week_off":
                    wo_count -= 1

        # Add adjusted counts
        for adj in adjustments:
            if adj.adjusted_status == "FD":
                fd_count += 1
            elif adj.adjusted_status == "HD":
                hd_count += 1
            elif adj.adjusted_status == "WO":
                wo_count += 1

        # 5. Approved leaves overlapping the cycle (pre-fetched)
        leaves = inputs.leaves

        paid_leave_count = Decimal("0.0")
        for leave in leaves:
            # Overlap days
            overlap_start = max(leave.start_date, cycle_from)
            overlap_end = min(leave.end_date, cycle_to)
            overlap_days = (overlap_end - overlap_start).days + 1
            if overlap_days > 0:
                paid_leave_count += Decimal(str(overlap_days))

        # 6. Resolve Paid Days
        # paid_day_count = present_days + half_days*0.5 + week_offs + paid_leaves
        off_day_basis = Decimal("0.0")
        if settings.off_day_compensation == "paid":
            off_day_basis = Decimal(str(wo_count)) * settings.off_day_wage_multiplier

        paid_day_count = Decimal(
            str(fd_count)) + (Decimal(str(hd_count)) * Decimal("0.5")
        ) + off_day_basis + paid_leave_count
        paid_day_count = min(paid_day_count, Decimal(str(total_days)))

        unpaid_day_count = Decimal(str(total_days)) - paid_day_count
        unpaid_day_count = max(Decimal("0.0"), unpaid_day_count)

        # 7. Derived Daily Wage
        if salary_type == "hourly":
            full_working_hours = settings.full_day_working_hours.hour + (
                settings.full_day_working_hours.minute / 60
            )
            daily_wage = monthly_salary * Decimal(str(full_working_hours))
        else:
            if settings.daily_wage_formula == "fixed_30":
                daily_wage = monthly_salary / Decimal("30.0")
            elif settings.daily_wage_formula == "fixed_26":
                daily_wage = monthly_salary / Decimal("26.0")
            else:  # calendar days
                daily_wage = monthly_salary / Decimal(str(total_days))

        daily_wage = daily_wage.quantize(Decimal("0.01"))

        # 8. Gross Wages Calculation
        if salary_type == "hourly":
            total_working_hours = Decimal(str(total_working_mins)) / Decimal("60.0")
            gross_wages = monthly_salary * total_working_hours
        else:
            gross_wages = daily_wage * paid_day_count

        gross_wages = gross_wages.quantize(Decimal("0.01"))

        # 9. Overtime Calculation
        overtime_amount = Decimal("0.00")
        if settings.overtime_hourly_multiplier > 0 and overtime_mins > 0:
            if salary_type == "hourly":
                hourly_rate = monthly_salary
            else:
                hourly_rate = daily_wage / Decimal("8.0")  # Assume standard 8-hour workday
            ot_hours = Decimal(str(overtime_mins)) / Decimal("60.0")
            overtime_amount = ot_hours * hourly_rate * settings.overtime_hourly_multiplier

        overtime_amount = overtime_amount.quantize(Decimal("0.01"))

        # 10. Penalties Calculation
        penalties_amount = Decimal("0.00")
        # Sum active (non-removed) manual penalties (pre-fetched)
        penalties_list = inputs.penalties
        for pen in penalties_list:
            penalties_amount += pen.penalty_amount

        # Automatic penalties
        if settings.full_day_penalty_enabled or settings.half_day_penalty_enabled:
            # Simple automatic penalty check
            pass

        penalties_amount = penalties_amount.quantize(Decimal("0.01"))

        # 11. Extra Hours Addition (pre-fetched)
        extras_amount = Decimal("0.00")
        extra_hours_records = inputs.extra_hours
        if salary_type == "hourly":
            hourly_rate = monthly_salary
        else:
            hourly_rate = daily_wage / Decimal("8.0")

        for record in extra_hours_records:
            extras_amount += record.extra_hours * hourly_rate

        extras_amount = extras_amount.quantize(Decimal("0.01"))

        # 12. Settlements - Loans/Advances installment deduction (pre-fetched)
        loan_advance_deduction = Decimal("0.00")
        for loan in inputs.loans:
            deductible = min(loan.monthly_installment, loan.outstanding_amount)
            loan_advance_deduction += deductible

        loan_advance_deduction = loan_advance_deduction.quantize(Decimal("0.01"))

        # 13. Settlements - Arrears addition (pre-fetched)
        arrears_amount = Decimal("0.00")
        arrears_rec = inputs.arrears
        if arrears_rec:
            arrears_amount = arrears_rec.outstanding_arrears

        arrears_amount = arrears_amount.quantize(Decimal("0.01"))

        # 14. Net pay arithmetic
        gross_earnings = gross_wages + overtime_amount + extras_amount
        gross_earnings = gross_earnings.quantize(Decimal("0.01"))

        net = gross_earnings + arrears_amount - penalties_amount - loan_advance_deduction
        net = net.quantize(Decimal("0.01"))

        if net < 0:
            to_pay = Decimal("0.00")
            balance_arrears = abs(net)
        else:
            to_pay = net
            balance_arrears = Decimal("0.00")

        return {
            "payroll_group_id": assignment.payroll_group_id,
            "employee_id": employee.employee_id,
            "cycle_from": cycle_from,
            "cycle_to": cycle_to,
            "total_days": total_days,
            "full_day_count": fd_count,
            "half_day_count": hd_count,
            "off_day_count": wo_count,
            "paid_leave_count": paid_leave_count,
            "paid_day_count": paid_day_count,
            "unpaid_day_count": unpaid_day_count,
            "daily_wage": daily_wage,
            "gross_wages": gross_wages,
            "overtime_amount": overtime_amount,
            "penalties_amount": penalties_amount,
            "extras_amount": extras_amount,
            "gross_earnings": gross_earnings,
            "loan_advance_deduction": loan_advance_deduction,
            "arrears_amount": arrears_amount,
            "to_pay": to_pay,
            "balance_arrears": balance_arrears,
            "payment_method": "bank_transfer",  # default fallback
            "is_finalized": False,
            "finalized_run_id": None,
            "computed_by": user_id,
            "computed_at": utcnow(),
        }

    # --- 7. Payroll Processing Operations ------------------------------------

    def _compute_batch(
        self,
        employees: list[Employee],
        payload: PayrollProcessRequestSchema,
        settings: PayrollSetting,
        inputs: PayrollBatchInputs,
        existing_rows: dict[int, PayrollComputedRow],
        user_id: int,
        *,
        skip_finalized: bool,
    ) -> tuple[list[PayrollProcessItemResultSchema], list[dict], list[dict]]:
        """Compute the whole batch in memory, returning results plus insert/update payloads.

        Performs no database access. ``skip_finalized`` reports an already-finalized
        employee as a ``PAYROLL_ALREADY_FINALIZED`` failure instead of recomputing it.
        """
        results: list[PayrollProcessItemResultSchema] = []
        to_insert: list[dict] = []
        to_update: list[dict] = []

        for emp in employees:
            existing = existing_rows.get(emp.employee_id)
            if skip_finalized and existing and existing.is_finalized:
                results.append(
                    PayrollProcessItemResultSchema(
                        employee_id=emp.employee_id,
                        success=False,
                        error_code="PAYROLL_ALREADY_FINALIZED",
                        error_message="Payroll already finalized for this employee in this "
                            "period.",
                    )
                )
                continue

            try:
                row_dict = self._calculate_employee_payroll(
                    emp,
                    payload.cycle_from,
                    payload.cycle_to,
                    settings,
                    user_id,
                    inputs.for_employee(emp.employee_id),
                )
            except Exception as e:
                results.append(
                    PayrollProcessItemResultSchema(
                        employee_id=emp.employee_id,
                        success=False,
                        error_code="VALIDATION_ERROR",
                        error_message=str(e),
                    )
                )
                continue

            if existing:
                to_update.append({"id": existing.id, **row_dict})
            else:
                to_insert.append(row_dict)

            results.append(
                PayrollProcessItemResultSchema(employee_id=emp.employee_id, success=True)
            )

        return results, to_insert, to_update

    async def _load_existing_rows(
        self, payload: PayrollProcessRequestSchema, employee_ids: list[int]
    ) -> dict[int, PayrollComputedRow]:
        """Fetch the cycle's existing computed rows for the whole employee set, keyed by id."""
        rows = await self.computed_rows.get_rows_for_cycle(
            payload.payroll_group_id, employee_ids, payload.cycle_from, payload.cycle_to
        )
        return {row.employee_id: row for row in rows}

    async def preview_payroll(
        self, org_id: int, payload: PayrollProcessRequestSchema
    ) -> list[PayrollComputedRow]:
        """Compute payroll preview records without persisting them to database."""
        await self._validate_payroll_group(org_id, payload.payroll_group_id)
        settings = await self._get_or_create_settings(org_id)

        employees = await self._resolve_employees(org_id, payload)
        employee_ids = [emp.employee_id for emp in employees]

        # Bulk-prefetch every calculator input once, then compute purely in memory.
        inputs = await self._prefetch_batch_inputs(
            org_id, employee_ids, payload.cycle_from, payload.cycle_to
        )

        preview_rows = []
        for emp in employees:
            try:
                row_dict = self._calculate_employee_payroll(
                    emp,
                    payload.cycle_from,
                    payload.cycle_to,
                    settings,
                    0,
                    inputs.for_employee(emp.employee_id),
                )
            except Exception:
                continue
            preview_rows.append(PayrollComputedRow(**row_dict))
        return preview_rows

    async def generate_payroll(
        self, org_id: int, payload: PayrollProcessRequestSchema, user_id: int
    ) -> PayrollProcessResponseSchema:
        """Compute and persist Draft payroll records, skipping already finalized periods."""
        await self._validate_payroll_group(org_id, payload.payroll_group_id)
        settings = await self._get_or_create_settings(org_id)

        employees = await self._resolve_employees(org_id, payload)
        employee_ids = [emp.employee_id for emp in employees]

        # --- Read phase: bulk-prefetch outside the transaction, so no row locks are held
        #     while we compute.
        inputs = await self._prefetch_batch_inputs(
            org_id, employee_ids, payload.cycle_from, payload.cycle_to
        )
        existing_rows = await self._load_existing_rows(payload, employee_ids)

        # --- Compute phase: pure, in memory, zero queries.
        results, to_insert, to_update = self._compute_batch(
            employees, payload, settings, inputs, existing_rows, user_id, skip_finalized=True
        )

        # --- Write phase: two bulk statements instead of one per employee.
        async with self.transaction():
            await self.computed_rows.bulk_insert_rows(to_insert)
            await self.computed_rows.bulk_update_rows(to_update)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="processing",
                action_type=ActionType.INSERT,
                title="Generate Payroll Run",
                description="Generated payroll calculation run for group ID "
                    f"{payload.payroll_group_id} from {payload.cycle_from} to {payload.cycle_to}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return PayrollProcessResponseSchema(results=results)

    async def recalculate_payroll(
        self, org_id: int, payload: PayrollProcessRequestSchema, user_id: int
    ) -> PayrollProcessResponseSchema:
        """Recalculate unfinalized records in the period, raising conflict if any is finalized."""
        await self._validate_payroll_group(org_id, payload.payroll_group_id)
        settings = await self._get_or_create_settings(org_id)

        employees = await self._resolve_employees(org_id, payload)
        employee_ids = [emp.employee_id for emp in employees]

        # --- Read phase. The existing rows are fetched ONCE and reused by both the
        #     finalized guard and the compute loop.
        inputs = await self._prefetch_batch_inputs(
            org_id, employee_ids, payload.cycle_from, payload.cycle_to
        )
        existing_rows = await self._load_existing_rows(payload, employee_ids)

        if any(row.is_finalized for row in existing_rows.values()):
            raise PayrollAlreadyFinalizedException()

        # --- Compute phase: pure, in memory, zero queries.
        results, to_insert, to_update = self._compute_batch(
            employees, payload, settings, inputs, existing_rows, user_id, skip_finalized=False
        )

        # --- Write phase.
        async with self.transaction():
            await self.computed_rows.bulk_insert_rows(to_insert)
            await self.computed_rows.bulk_update_rows(to_update)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="processing",
                action_type=ActionType.UPDATE,
                title="Recalculate Payroll Run",
                description="Recalculated payroll calculation run for group ID "
                    f"{payload.payroll_group_id} from {payload.cycle_from} to {payload.cycle_to}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return PayrollProcessResponseSchema(results=results)

    async def finalize_payroll(
        self, org_id: int, payload: PayrollProcessRequestSchema, user_id: int
    ) -> FinalizedPayrollRun:
        """Lock and finalize payroll, recording loans and arrears ledger transactions."""
        await self._validate_payroll_group(org_id, payload.payroll_group_id)
        settings = await self._get_or_create_settings(org_id)

        # Check existing run
        stmt_run = select(FinalizedPayrollRun).where(
            FinalizedPayrollRun.org_id == org_id,
            FinalizedPayrollRun.payroll_group_id == payload.payroll_group_id,
            FinalizedPayrollRun.cycle_from == payload.cycle_from,
            FinalizedPayrollRun.cycle_to == payload.cycle_to,
            FinalizedPayrollRun.is_definalized.is_(False),
        )
        existing_run = (await self.session.execute(stmt_run)).scalar_one_or_none()
        if existing_run:
            raise PayrollAlreadyFinalizedException()

        # Query all computed rows in range
        stmt_rows = select(PayrollComputedRow).where(
            PayrollComputedRow.payroll_group_id == payload.payroll_group_id,
            PayrollComputedRow.cycle_from == payload.cycle_from,
            PayrollComputedRow.cycle_to == payload.cycle_to,
        )
        comp_rows = (await self.session.execute(stmt_rows)).scalars().all()
        if not comp_rows:
            raise ConflictException("No payroll records generated to finalize.")

        total_amount = Decimal("0.00")
        for row in comp_rows:
            if row.is_finalized:
                raise PayrollAlreadyFinalizedException()
            if row.to_pay < 0 and not settings.full_day_penalty_enabled:  # fallback check
                raise ValidationException(f"Employee {row.employee_id} has negative net pay.")
            total_amount += row.to_pay

        async with self.transaction():
            # 1. Create run record
            run = await self.runs.create(
                {
                    "org_id": org_id,
                    "payroll_group_id": payload.payroll_group_id,
                    "cycle_from": payload.cycle_from,
                    "cycle_to": payload.cycle_to,
                    "payroll_module": "core_payroll",
                    "finalized_amount": total_amount,
                    "finalized_at": utcnow(),
                    "finalized_by": user_id,
                    "payment_status": PaymentStatus.PENDING.value,
                    "is_definalized": False,
                }
            )

            # 1b. Create corresponding PayrollFinalization record for history details page
            gross_sum = sum(Decimal(str(r.gross_earnings or 0)) for r in comp_rows)
            deduction_sum = sum(Decimal(str(r.total_deductions or 0)) for r in comp_rows)
            fin_obj = PayrollFinalization(
                org_id=org_id,
                payroll_group_id=payload.payroll_group_id,
                from_date=payload.cycle_from,
                to_date=payload.cycle_to,
                payroll_module="Monthly Payroll",
                employee_count=len(comp_rows),
                gross_amount=gross_sum,
                deduction_amount=deduction_sum,
                net_payable=total_amount,
                finalized_amount=total_amount,
                status="Finalized",
                finalized_by=user_id,
                finalized_on=utcnow(),
            )
            self.session.add(fin_obj)
            await self.session.flush()

            for r in comp_rows:
                emp_obj = PayrollFinalizationEmployee(
                    payroll_finalization_id=fin_obj.id,
                    employee_id=r.employee_id,
                    loan_amount=Decimal(str(r.loan_advance_deduction or 0)),
                    arrears_amount=Decimal(str(r.arrears_amount or 0)),
                    net_salary=Decimal(str(r.to_pay or 0)),
                    json_snapshot={
                        "employee_id": r.employee_id,
                        "gross_earnings": str(r.gross_earnings or 0),
                        "to_pay": str(r.to_pay or 0),
                    },
                )
                self.session.add(emp_obj)

            # 2. Lock computed rows
            for row in comp_rows:
                await self.computed_rows.update(
                    row, {"is_finalized": True, "finalized_run_id": run.id}
                )

                # 3. Process settlements (Loans)
                if row.loan_advance_deduction > 0:
                    deducted_left = row.loan_advance_deduction
                    stmt_loans = select(EmployeeLoanAdvance).where(
                        EmployeeLoanAdvance.org_id == org_id,
                        EmployeeLoanAdvance.employee_id == row.employee_id,
                        EmployeeLoanAdvance.status == "active",
                    ).order_by(EmployeeLoanAdvance.id)
                    loans = (await self.session.execute(stmt_loans)).scalars().all()
                    for loan in loans:
                        if deducted_left <= 0:
                            break
                        installment = min(loan.monthly_installment, loan.outstanding_amount)
                        deduction = min(installment, deducted_left)
                        deducted_left -= deduction

                        new_outstanding = loan.outstanding_amount - deduction
                        await self.session.execute(
                            select(EmployeeLoanAdvance).where(EmployeeLoanAdvance.id == loan.id)
                        )  # load into session
                        loan.outstanding_amount = new_outstanding
                        if new_outstanding <= 0:
                            loan.status = "closed"

                        # Create loan transaction row
                        tx = LoanAdvanceTransaction(
                            org_id=org_id,
                            loan_advance_id=loan.id,
                            employee_id=row.employee_id,
                            transaction_date=utcnow().date(),
                            transaction_type="credit",
                            amount=deduction,
                            installment_amount=deduction,
                            type_label=loan.type,
                            source="payroll",
                            payroll_run_id=run.id,
                            created_by=user_id,
                        )
                        self.session.add(tx)

                # 4. Process settlements (Arrears)
                if row.arrears_amount > 0:
                    stmt_arr = select(EmployeeArrears).where(
                        EmployeeArrears.org_id == org_id,
                        EmployeeArrears.employee_id == row.employee_id,
                    )
                    arrears_rec = (await self.session.execute(stmt_arr)).scalar_one_or_none()
                    if arrears_rec:
                        outstanding_before = arrears_rec.outstanding_arrears
                        new_paid = arrears_rec.arrears_paid + row.arrears_amount
                        new_outstanding = arrears_rec.outstanding_arrears - row.arrears_amount
                        arrears_rec.arrears_paid = new_paid
                        arrears_rec.outstanding_arrears = new_outstanding

                        # Create arrears transaction row
                        tx = ArrearsTransaction(
                            org_id=org_id,
                            employee_arrears_id=arrears_rec.id,
                            employee_id=row.employee_id,
                            transaction_date=utcnow().date(),
                            transaction_type="credit",
                            amount=row.arrears_amount,
                            outstanding_before=outstanding_before,
                            outstanding_after=new_outstanding,
                            source="payroll",
                            payroll_run_id=run.id,
                            created_by=user_id,
                        )
                        self.session.add(tx)

            # 5. Mark cycle as finalized
            cycle = await self.cycles.get_cycle(payload.payroll_group_id, payload.cycle_to)
            if cycle:
                await self.cycles.update(cycle, {"is_finalized": True})

            # 6. Notify affected employees' linked users (one query, one
            #    multi-recipient notification; unlinked employees are skipped).
            notifier = self._get_notifier()
            recipient_user_ids = await notifier.resolve_user_ids_for_employees(
                org_id, sorted({row.employee_id for row in comp_rows})
            )
            await notifier.emit_system_notification(
                org_id,
                recipient_user_ids=recipient_user_ids,
                title="Payroll Finalized",
                message=(
                    f"Your payroll for {payload.cycle_from} to {payload.cycle_to} "
                    "has been finalized."
                ),
                notification_type=NotificationType.PAYROLL.value,
                source_module="payroll",
                source_entity_type="finalized_payroll_run",
                source_entity_id=run.id,
                created_by=user_id,
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="processing",
                action_type=ActionType.UPDATE,
                title="Finalize Payroll Run",
                description=f"Finalized & locked payroll run ID {run.id} for amount "
                    f"{total_amount}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return run

    async def definalize_payroll(
        self, org_id: int, run_id: int, user_id: int
    ) -> FinalizedPayrollRun:
        """Unlock payroll run, reversing loan adjustments and arrears changes."""
        run = await self.runs.get_by_id_in_org(org_id, run_id)
        if not run:
            raise FinalizedRunNotFoundException()
        if run.is_definalized:
            raise PayrollNotFinalizedException("Payroll run is already unlocked.")

        async with self.transaction():
            # 1. Reverse Loan transactions
            stmt_loan_txs = select(LoanAdvanceTransaction).where(
                LoanAdvanceTransaction.payroll_run_id == run_id
            )
            loan_txs = (await self.session.execute(stmt_loan_txs)).scalars().all()
            for tx in loan_txs:
                # restore loan balance
                stmt_loan = select(
                    EmployeeLoanAdvance).where(EmployeeLoanAdvance.id == tx.loan_advance_id
                )
                loan = (await self.session.execute(stmt_loan)).scalar_one_or_none()
                if loan:
                    loan.outstanding_amount += tx.amount
                    if loan.outstanding_amount > 0:
                        loan.status = "active"
                await self.session.delete(tx)

            # 2. Reverse Arrears transactions
            stmt_arr_txs = select(ArrearsTransaction).where(
                ArrearsTransaction.payroll_run_id == run_id
            )
            arr_txs = (await self.session.execute(stmt_arr_txs)).scalars().all()
            for tx in arr_txs:
                # restore arrears balance
                stmt_arr = select(
                    EmployeeArrears).where(EmployeeArrears.id == tx.employee_arrears_id
                )
                arr = (await self.session.execute(stmt_arr)).scalar_one_or_none()
                if arr:
                    arr.outstanding_arrears += tx.amount
                    arr.arrears_paid -= tx.amount
                await self.session.delete(tx)

            # 3. Unlock computed rows
            stmt_rows = select(PayrollComputedRow).where(
                PayrollComputedRow.finalized_run_id == run_id
            )
            comp_rows = (await self.session.execute(stmt_rows)).scalars().all()
            for row in comp_rows:
                await self.computed_rows.update(
                    row, {"is_finalized": False, "finalized_run_id": None}
                )

            # 4. Unfinalize Cycle
            cycle = await self.cycles.get_cycle(run.payroll_group_id, run.cycle_to)
            if cycle:
                await self.cycles.update(cycle, {"is_finalized": False})

            # 5. Mark run as definalized
            updated_run = await self.runs.update(
                run,
                {
                    "is_definalized": True,
                    "definalized_at": utcnow(),
                    "definalized_by": user_id,
                },
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="processing",
                action_type=ActionType.UPDATE,
                title="Unlock Payroll Run",
                description=f"Unlocked payroll run ID {run_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return updated_run

    async def record_payment(
        self, org_id: int, run_id: int, payload: RecordPaymentRequestSchema, user_id: int
    ) -> FinalizedPayrollRun:
        """Record disbursement parameters against a finalized lock run."""
        run = await self.runs.get_by_id_in_org(org_id, run_id)
        if not run:
            raise FinalizedRunNotFoundException()

        async with self.transaction():
            updated = await self.runs.update(
                run,
                {
                    "paid_amount": payload.paid_amount,
                    "paid_at": payload.paid_at or utcnow(),
                    "payment_status": payload.payment_status.value,
                },
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="processing",
                action_type=ActionType.UPDATE,
                title="Record Payroll Payment",
                description=f"Recorded payment for run ID {run_id} status "
                    f"{payload.payment_status}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return updated

    async def list_finalized_runs(
        self,
        org_id: int,
        group_id: int | None,
        cycle_from: date | None,
        cycle_to: date | None,
        payment_status: PaymentStatus | None,
        page: int,
        page_size: int,
    ) -> PaginatedResponse[FinalizedPayrollRun]:
        """List paginated finalized lock runs."""
        runs = await self.runs.search(
            org_id,
            payroll_group_id=group_id,
            cycle_from=cycle_from,
            cycle_to=cycle_to,
            payment_status=payment_status,
            page=page,
            page_size=page_size,
        )
        total = await self.runs.search_count(
            org_id,
            payroll_group_id=group_id,
            cycle_from=cycle_from,
            cycle_to=cycle_to,
            payment_status=payment_status,
        )
        return self.paginate(runs, page=page, page_size=page_size, total_records=total)

    async def get_finalized_run(self, org_id: int, run_id: int) -> FinalizedPayrollRun:
        """Get details of a finalized lock run."""
        run = await self.runs.get_by_id_in_org(org_id, run_id)
        if not run:
            raise FinalizedRunNotFoundException()
        return run

    # --- 8. Payroll Records & Summary ---------------------------------------

    async def list_records(
        self,
        org_id: int,
        group_id: int | None,
        cycle_from: date | None,
        cycle_to: date | None,
        employee_id: int | None,
        is_finalized: bool | None,
        branch_id: int | None,
        dept_id: int | None,
        page: int,
        page_size: int,
    ) -> PaginatedResponse[PayrollComputedRow]:
        """List paginated computed payroll records (supports branch/dept scoping)."""
        rows = await self.computed_rows.search(
            org_id,
            payroll_group_id=group_id,
            cycle_from=cycle_from,
            cycle_to=cycle_to,
            employee_id=employee_id,
            is_finalized=is_finalized,
            branch_id=branch_id,
            dept_id=dept_id,
            page=page,
            page_size=page_size,
        )
        total = await self.computed_rows.search_count(
            org_id,
            payroll_group_id=group_id,
            cycle_from=cycle_from,
            cycle_to=cycle_to,
            employee_id=employee_id,
            is_finalized=is_finalized,
            branch_id=branch_id,
            dept_id=dept_id,
        )
        return self.paginate(rows, page=page, page_size=page_size, total_records=total)

    async def get_record(self, org_id: int, row_id: int) -> PayrollComputedRow:
        """Retrieve a specific computed payroll row."""
        row = await self.computed_rows.get_by_id(row_id)
        if not row:
            raise ComputedRowNotFoundException()

        # org validation
        await self._validate_payroll_group(org_id, row.payroll_group_id)
        return row

    async def get_summary(
        self, org_id: int, group_id: int, cycle_from: date, cycle_to: date
    ) -> PayrollSummaryResponseSchema:
        """Aggregate summary statistics over computed records."""
        await self._validate_payroll_group(org_id, group_id)
        metrics = await self.computed_rows.get_summary(org_id, group_id, cycle_from, cycle_to)
        return PayrollSummaryResponseSchema(
            headcount=metrics["headcount"],
            total_gross_earnings=metrics["total_gross_earnings"] or Decimal("0.00"),
            total_to_pay=metrics["total_to_pay"] or Decimal("0.00"),
            total_overtime=metrics["total_overtime"] or Decimal("0.00"),
            total_penalties=metrics["total_penalties"] or Decimal("0.00"),
            total_deductions=metrics["total_deductions"] or Decimal("0.00"),
        )

    async def get_employee_history(
        self, org_id: int, employee_id: int, page: int, page_size: int
    ) -> PaginatedResponse[PayrollComputedRow]:
        """Fetch historical payslips / computed rows of a specific employee."""
        await self._validate_employee(org_id, employee_id)
        rows = await self.computed_rows.get_employee_history(
            employee_id, page=page, page_size=page_size
        )
        total = await self.computed_rows.get_employee_history_count(employee_id)
        return self.paginate(rows, page=page, page_size=page_size, total_records=total)

    # --- 9. Payslips ---------------------------------------------------------

    async def view_payslip(self, org_id: int, row_id: int) -> PayslipResponseSchema:
        """Render payslip structure details on demand from calculated data."""
        row = await self.get_record(org_id, row_id)
        employee = await self.employees.get_by_id(row.employee_id)
        emp_name = employee.employee_name if employee else None
        emp_code = employee.employee_code if employee else None

        # Build structures
        earnings = [
            PayslipSectionItemSchema(
                key="gross_wages", label="Gross Wages / Basic", value=row.gross_wages
            ),
            PayslipSectionItemSchema(
                key="overtime", label="Overtime Earnings", value=row.overtime_amount
            ),
            PayslipSectionItemSchema(
                key="extras", label="Extra Hours / Adjustments Additions", value=row.extras_amount
            ),
            PayslipSectionItemSchema(key="arrears", label="Arrears", value=row.arrears_amount),
        ]
        deductions = [
            PayslipSectionItemSchema(
                key="penalties", label="Manual Penalties / Deductions", value=row.penalties_amount
            ),
            PayslipSectionItemSchema(
                key="loans", label="Loan / Advance Installment", value=row.loan_advance_deduction
            ),
        ]

        return PayslipResponseSchema(
            row_id=row.id,
            employee_id=row.employee_id,
            employee_name=emp_name,
            employee_code=emp_code,
            cycle_from=row.cycle_from,
            cycle_to=row.cycle_to,
            earnings=earnings,
            deductions=deductions,
            net_pay=row.to_pay,
            payment_method=row.payment_method,
            is_finalized=row.is_finalized,
        )

    async def download_payslip_pdf(self, org_id: int, row_id: int) -> bytes:
        """Download payslip PDF binary stream (only allowed for finalized records)."""
        row = await self.get_record(org_id, row_id)
        if not row.is_finalized:
            raise ConflictException("Payslip PDF is only available for finalized payroll records.")

        # Simulate rendering of PDF bytes
        pdf_buffer = io.BytesIO()
        pdf_buffer.write(b"%PDF-1.4 mock payslip bytes for computed row ")
        pdf_buffer.write(str(row_id).encode())
        return pdf_buffer.getvalue()

    # --- 10. Attendance Adjustments ------------------------------------------

    async def add_adjustment(
        self, org_id: int, payload: AttendanceAdjustmentCreateSchema, user_id: int
    ) -> AttendanceAdjustment:
        """Create or update attendance status overrides, verifying period is not finalized."""
        await self._validate_employee(org_id, payload.employee_id)

        # Check if period finalized
        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id == payload.employee_id,
            PayrollComputedRow.cycle_from <= payload.attendance_date,
            PayrollComputedRow.cycle_to >= payload.attendance_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        async with self.transaction():
            existing = await self.adjustments.get_adjustment(
                payload.employee_id, payload.attendance_date
            )
            if existing:
                if not payload.is_forced_overwrite:
                    raise AdjustmentExistsException()
                # Update existing
                adj = await self.adjustments.update(
                    existing,
                    {
                        "adjusted_status": payload.adjusted_status.value,
                        "original_status": (
                            payload.original_status.value if payload.original_status else None
                        ),
                        "has_punch_error": payload.has_punch_error,
                        "adjustment_source": payload.adjustment_source.value,
                        "adjusted_by": user_id,
                        "adjusted_at": utcnow(),
                    },
                )
            else:
                adj = await self.adjustments.create(
                    {
                        "org_id": org_id,
                        "employee_id": payload.employee_id,
                        "attendance_date": payload.attendance_date,
                        "original_status": (
                            payload.original_status.value if payload.original_status else None
                        ),
                        "adjusted_status": payload.adjusted_status.value,
                        "is_forced_overwrite": payload.is_forced_overwrite,
                        "has_punch_error": payload.has_punch_error,
                        "adjustment_source": payload.adjustment_source.value,
                        "adjusted_by": user_id,
                        "adjusted_at": utcnow(),
                    }
                )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="adjustments",
                action_type=ActionType.INSERT,
                title="Add Attendance Adjustment",
                description=f"Added/updated attendance override for employee {payload.employee_id} "
                    f"on {payload.attendance_date}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=payload.employee_id,
            )
        return adj

    async def update_adjustment(
        self,
        org_id: int,
        adjustment_id: int,
        payload: AttendanceAdjustmentUpdateSchema,
        user_id: int,
    ) -> AttendanceAdjustment:
        """Modify an override record, verifying finalization lock constraints."""
        adj = await self.adjustments.get_by_id(adjustment_id)
        if not adj or adj.org_id != org_id:
            raise AdjustmentNotFoundException()

        # Check if period finalized
        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id == adj.employee_id,
            PayrollComputedRow.cycle_from <= adj.attendance_date,
            PayrollComputedRow.cycle_to >= adj.attendance_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        update_data = payload.model_dump(exclude_unset=True)
        if "adjusted_status" in update_data:
            update_data["adjusted_status"] = update_data["adjusted_status"].value
        if "original_status" in update_data:
            update_data["original_status"] = update_data["original_status"].value

        async with self.transaction():
            update_data["adjusted_by"] = user_id
            update_data["adjusted_at"] = utcnow()
            updated = await self.adjustments.update(adj, update_data)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="adjustments",
                action_type=ActionType.UPDATE,
                title="Update Attendance Adjustment",
                description=f"Updated attendance override ID {adjustment_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=adj.employee_id,
            )
        return updated

    async def delete_adjustment(self, org_id: int, adjustment_id: int, user_id: int) -> None:
        """Remove override records, verifying finalization lock constraints."""
        adj = await self.adjustments.get_by_id(adjustment_id)
        if not adj or adj.org_id != org_id:
            raise AdjustmentNotFoundException()

        # Check if period finalized
        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id == adj.employee_id,
            PayrollComputedRow.cycle_from <= adj.attendance_date,
            PayrollComputedRow.cycle_to >= adj.attendance_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        async with self.transaction():
            await self.adjustments.delete(adj)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="adjustments",
                action_type=ActionType.DELETE,
                title="Delete Attendance Adjustment",
                description=f"Removed attendance override ID {adjustment_id}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=adj.employee_id,
            )

    async def list_adjustments(
        self,
        org_id: int,
        employee_id: int | None,
        date_from: date | None,
        date_to: date | None,
        page: int,
        page_size: int,
    ) -> PaginatedResponse[AttendanceAdjustment]:
        """List paginated adjustments (supports branch/dept scoping)."""
        adjustments = await self.adjustments.search(
            org_id,
            employee_id=employee_id,
            date_from=date_from,
            date_to=date_to,
            page=page,
            page_size=page_size,
        )
        total = await self.adjustments.search_count(
            org_id,
            employee_id=employee_id,
            date_from=date_from,
            date_to=date_to,
        )
        return self.paginate(adjustments, page=page, page_size=page_size, total_records=total)

    async def add_penalty(
        self, org_id: int, payload: AttendanceAdjustmentPenaltyCreateSchema, user_id: int
    ) -> AttendanceAdjustmentPenalty:
        """Create custom penalties mapping to salary deductions, checking finalization status."""
        await self._validate_employee(org_id, payload.employee_id)

        # Check if period finalized
        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id == payload.employee_id,
            PayrollComputedRow.cycle_from <= payload.attendance_date,
            PayrollComputedRow.cycle_to >= payload.attendance_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        async with self.transaction():
            penalty = await self.penalties.create(
                {
                    "employee_id": payload.employee_id,
                    "attendance_date": payload.attendance_date,
                    "penalty_amount": payload.penalty_amount,
                    "remark": payload.remark,
                    "is_removed": False,
                    "created_by": user_id,
                }
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="adjustments",
                action_type=ActionType.INSERT,
                title="Add Attendance Penalty",
                description=f"Recorded penalty of {payload.penalty_amount} for employee "
                    f"{payload.employee_id} on {payload.attendance_date}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=payload.employee_id,
            )
        return penalty

    async def add_extra_hours(
        self, org_id: int, payload: AttendanceAdjustmentExtraHoursCreateSchema, user_id: int
    ) -> AttendanceAdjustmentExtraHours:
        """Create custom extra hours logs mapping to salary additions, checking finalization status.
        """
        await self._validate_employee(org_id, payload.employee_id)

        # Check if period finalized
        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id == payload.employee_id,
            PayrollComputedRow.cycle_from <= payload.attendance_date,
            PayrollComputedRow.cycle_to >= payload.attendance_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        async with self.transaction():
            existing = await self.extra_hours.get_extra_hours(
                payload.employee_id, payload.attendance_date
            )
            if existing:
                raise AdjustmentExistsException("Extra hours log already exists for this date.")

            eh = await self.extra_hours.create(
                {
                    "employee_id": payload.employee_id,
                    "attendance_date": payload.attendance_date,
                    "extra_hours": payload.extra_hours,
                    "remark": payload.remark,
                    "created_by": user_id,
                }
            )

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="adjustments",
                action_type=ActionType.INSERT,
                title="Add Extra Hours Adjustment",
                description=f"Recorded {payload.extra_hours} extra hours for employee "
                    f"{payload.employee_id} on {payload.attendance_date}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
                employee_id=payload.employee_id,
            )
        return eh

    # ===========================================================================
    # Bulk Attendance Adjustments Matrix (Phase 2)
    # ===========================================================================

    async def get_bulk_attendance_matrix(
        self,
        org_id: int,
        date_from: date,
        date_to: date,
        branch_id: int | None = None,
        dept_id: int | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 25,
    ) -> BulkAttendanceAdjustmentMatrixResponseSchema:
        """Fetch attendance status matrix for employees over a date range."""
        employees = await self.adjustments.search_matrix_employees(
            org_id=org_id,
            branch_id=branch_id,
            dept_id=dept_id,
            search=search,
            page=page,
            page_size=page_size,
        )
        total_records = await self.adjustments.count_matrix_employees(
            org_id=org_id,
            branch_id=branch_id,
            dept_id=dept_id,
            search=search,
        )

        date_strs: list[str] = []
        curr = date_from
        while curr <= date_to and len(date_strs) < 31:
            date_strs.append(curr.isoformat())
            curr = curr + timedelta(days=1)

        if not employees:
            meta = PaginationMeta.build(page=page, page_size=page_size, total_records=total_records)
            return BulkAttendanceAdjustmentMatrixResponseSchema(
                dates=date_strs,
                items=[],
                pagination=meta,
            )

        emp_ids = [e.employee_id for e in employees]

        adjustments_list = await self.adjustments.get_adjustments_for_employees(
            org_id=org_id,
            employee_ids=emp_ids,
            date_from=date_from,
            date_to=date_to,
        )
        adj_map: dict[tuple[int, str], str] = {
            (a.employee_id, a.attendance_date.isoformat()): a.adjusted_status for a in adjustments_list
        }

        items: list[BulkAttendanceAdjustmentMatrixItemSchema] = []
        for emp in employees:
            attendance_map: dict[str, str] = {}
            for d_str in date_strs:
                if (emp.employee_id, d_str) in adj_map:
                    attendance_map[d_str] = adj_map[(emp.employee_id, d_str)]
                else:
                    d_obj = date.fromisoformat(d_str)
                    attendance_map[d_str] = "WO" if d_obj.weekday() == 6 else "FD"

            items.append(
                BulkAttendanceAdjustmentMatrixItemSchema(
                    employee_id=emp.employee_id,
                    employee_code=emp.employee_code,
                    employee_name=emp.employee_name,
                    department_name=getattr(emp, "department_name", None) or "General",
                    designation_name=getattr(emp, "designation_name", None) or "Staff",
                    branch_id=emp.master_branch_id,
                    branch_name=getattr(emp, "branch_name", None) or "Main HQ",
                    attendance=attendance_map,
                )
            )

        meta = PaginationMeta.build(page=page, page_size=page_size, total_records=total_records)
        return BulkAttendanceAdjustmentMatrixResponseSchema(
            dates=date_strs,
            items=items,
            pagination=meta,
        )

    async def batch_update_bulk_attendance_adjustments(
        self,
        org_id: int,
        payload: BulkAttendanceAdjustmentBatchUpdateSchema,
        user_id: int,
    ) -> BulkAttendanceAdjustmentBatchUpdateResponseSchema:
        """Batch save modified attendance cells for employees."""
        if not payload.updates:
            return BulkAttendanceAdjustmentBatchUpdateResponseSchema(
                updated_count=0,
                message="No cell updates provided.",
            )

        emp_ids = list({u.employee_id for u in payload.updates})
        min_date = min(u.attendance_date for u in payload.updates)
        max_date = max(u.attendance_date for u in payload.updates)

        stmt_comp = select(PayrollComputedRow.id).where(
            PayrollComputedRow.employee_id.in_(emp_ids),
            PayrollComputedRow.cycle_from <= max_date,
            PayrollComputedRow.cycle_to >= min_date,
            PayrollComputedRow.is_finalized.is_(True),
        )
        finalized = (await self.session.execute(stmt_comp.limit(1))).first() is not None
        if finalized:
            raise PayrollAlreadyFinalizedException()

        saved_count = 0
        async with self.transaction():
            for update_item in payload.updates:
                existing = await self.adjustments.get_adjustment(
                    update_item.employee_id, update_item.attendance_date
                )
                if existing:
                    await self.adjustments.update(
                        existing,
                        {
                            "adjusted_status": update_item.adjusted_status,
                            "original_status": update_item.original_status,
                            "adjusted_by": user_id,
                            "adjusted_at": utcnow(),
                            "adjustment_source": "spreadsheet",
                        },
                    )
                else:
                    await self.adjustments.create(
                        {
                            "org_id": org_id,
                            "employee_id": update_item.employee_id,
                            "attendance_date": update_item.attendance_date,
                            "original_status": update_item.original_status,
                            "adjusted_status": update_item.adjusted_status,
                            "is_forced_overwrite": True,
                            "has_punch_error": False,
                            "adjustment_source": "spreadsheet",
                            "adjusted_by": user_id,
                            "adjusted_at": utcnow(),
                        }
                    )
                saved_count += 1

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="bulk_adjustments",
                action_type=ActionType.UPDATE,
                title="Batch Bulk Attendance Adjustments",
                description=f"Saved batch update of {saved_count} attendance status cell adjustments.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        return BulkAttendanceAdjustmentBatchUpdateResponseSchema(
            updated_count=saved_count,
            message=f"Successfully saved {saved_count} attendance adjustments.",
        )

    async def reset_bulk_attendance_adjustments(
        self,
        org_id: int,
        payload: BulkAttendanceAdjustmentResetSchema,
        user_id: int,
    ) -> int:
        """Reset bulk attendance adjustments for a given date range and scope."""
        async with self.transaction():
            reset_count = await self.adjustments.bulk_reset_adjustments(
                org_id=org_id,
                date_from=payload.date_from,
                date_to=payload.date_to,
                branch_id=payload.branch_id,
                employee_ids=payload.employee_ids,
            )
            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="bulk_adjustments",
                action_type=ActionType.DELETE,
                title="Reset Bulk Attendance Adjustments",
                description=f"Reset {reset_count} attendance adjustments between {payload.date_from} and {payload.date_to}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )
        return reset_count

    async def export_bulk_attendance_adjustments_excel(
        self,
        org_id: int,
        date_from: date,
        date_to: date,
        branch_id: int | None = None,
        dept_id: int | None = None,
        search: str | None = None,
    ) -> bytes:
        """Export attendance adjustment matrix as an Excel workbook (.xlsx)."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        matrix_data = await self.get_bulk_attendance_matrix(
            org_id=org_id,
            date_from=date_from,
            date_to=date_to,
            branch_id=branch_id,
            dept_id=dept_id,
            search=search,
            page=1,
            page_size=10000,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bulk Attendance Adjustments"

        headers = ["Employee ID", "Employee Name", "Department", "Designation", "Branch"]
        for d_str in matrix_data.dates:
            d_obj = date.fromisoformat(d_str)
            headers.append(f"{d_obj.strftime('%d-%b')} ({d_obj.strftime('%a')})")

        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for emp in matrix_data.items:
            row = [
                emp.employee_code,
                emp.employee_name,
                emp.department_name,
                emp.designation_name,
                emp.branch_name,
            ]
            for d_str in matrix_data.dates:
                row.append(emp.attendance.get(d_str, "FD"))
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # --- 9. Process Payroll APIs (Phase 2) -----------------------------------

    async def get_process_payroll_matrix(
        self,
        org_id: int,
        date_from: date,
        date_to: date,
        payroll_group_id: int | None = None,
        branch_id: int | None = None,
        dept_id: int | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 25,
    ) -> PayrollRecordListResponse:
        """Fetch process payroll matrix for employees over a date range."""
        rows = await self.computed_rows.search(
            org_id=org_id,
            payroll_group_id=payroll_group_id,
            cycle_from=date_from,
            cycle_to=date_to,
            branch_id=branch_id,
            dept_id=dept_id,
            page=page,
            page_size=page_size,
        )
        total_records = await self.computed_rows.search_count(
            org_id=org_id,
            payroll_group_id=payroll_group_id,
            cycle_from=date_from,
            cycle_to=date_to,
            branch_id=branch_id,
            dept_id=dept_id,
        )
        meta = PaginationMeta.build(page=page, page_size=page_size, total_records=total_records)
        items = [PayrollComputedRowSchema.model_validate(row) for row in rows]
        return PayrollRecordListResponse(items=items, pagination=meta)

    async def calculate_process_payroll(
        self,
        org_id: int,
        payload: PayrollProcessRequestSchema,
        user_id: int,
    ) -> PayrollProcessResponseSchema:
        """Calculate or recalculate payroll matrix for target employees."""
        return await self.generate_payroll_records(org_id=org_id, payload=payload, user_id=user_id)

    async def finalize_process_payroll(
        self,
        org_id: int,
        payload: PayrollProcessRequestSchema,
        user_id: int,
    ) -> FinalizedPayrollRunResponseSchema:
        """Lock and finalize payroll run for a period."""
        return await self.finalize_payroll(org_id=org_id, payload=payload, user_id=user_id)

    async def export_process_payroll(
        self,
        org_id: int,
        date_from: date,
        date_to: date,
        payroll_group_id: int | None = None,
        branch_id: int | None = None,
        dept_id: int | None = None,
    ) -> bytes:
        """Generate Excel spreadsheet for process payroll matrix."""
        rows = await self.computed_rows.search(
            org_id=org_id,
            payroll_group_id=payroll_group_id,
            cycle_from=date_from,
            cycle_to=date_to,
            branch_id=branch_id,
            dept_id=dept_id,
            page=1,
            page_size=1000,
        )

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Process Payroll"

        headers = [
            "Employee ID", "Full Day", "Half Day", "Off Days", "Paid Leaves",
            "Paid Days", "Unpaid Days", "Daily Wage", "Gross Wages",
            "Overtime", "Penalties", "Extras", "Gross Earnings",
            "Loan & Advance", "Arrears", "Net Payable",
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0B85C9", end_color="0B85C9", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for r in rows:
            ws.append([
                r.employee_id,
                r.full_day_count,
                r.half_day_count,
                r.off_day_count,
                float(r.paid_leave_count),
                float(r.paid_day_count),
                float(r.unpaid_day_count),
                float(r.daily_wage),
                float(r.gross_wages),
                float(r.overtime_amount),
                float(r.penalties_amount),
                float(r.extras_amount),
                float(r.gross_earnings),
                float(r.loan_advance_deduction),
                float(r.arrears_amount),
                float(r.to_pay),
            ])

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    async def get_process_payroll_employee_detail(
        self,
        org_id: int,
        employee_id: int,
        date_from: date,
        date_to: date,
    ) -> PayrollComputedRowSchema:
        """Fetch detailed computed row breakdown for a single employee."""
        rows = await self.computed_rows.search(
            org_id=org_id,
            employee_id=employee_id,
            cycle_from=date_from,
            cycle_to=date_to,
            page=1,
            page_size=1,
        )
        if not rows:
            raise NotFoundException(f"Payroll record for employee {employee_id} not found.")
        return PayrollComputedRowSchema.model_validate(rows[0])

    # ===========================================================================
    # 20. Finalized Payroll History & Immutability Engine
    # ===========================================================================

    async def finalize_payroll_record(
        self,
        org_id: int,
        payload: PayrollFinalizationCreateSchema,
        user_id: int,
    ) -> PayrollFinalizationResponseSchema:
        """Finalize payroll for a group & cycle range, creating an immutable history record and frozen employee snapshots."""
        group = await self._validate_payroll_group(org_id, payload.payroll_group_id)

        # Process/compute rows for group employees in date range
        process_resp = await self.process_payroll(
            org_id=org_id,
            payload=PayrollProcessRequestSchema(
                payroll_group_id=payload.payroll_group_id,
                date_from=payload.from_date,
                date_to=payload.to_date,
                salary_type="monthly",
            ),
            user_id=user_id,
        )

        computed_items = process_resp.items
        emp_ids = [r.employee_id for r in computed_items]
        emp_map: dict[int, Employee] = {}
        if emp_ids:
            emp_rows = await self.employees.get_by_ids(org_id, emp_ids)
            emp_map = {e.employee_id: e for e in emp_rows}

        gross_sum = Decimal("0.00")
        deduction_sum = Decimal("0.00")
        net_sum = Decimal("0.00")

        employee_snapshots = []
        for r in computed_items:
            emp_info = emp_map.get(r.employee_id)
            emp_code = emp_info.employee_code if emp_info else f"EMP{r.employee_id}"
            emp_name = emp_info.employee_name if emp_info else f"Employee {r.employee_id}"

            gross = Decimal(str(r.gross_earnings))
            deduction = Decimal(str(r.penalties_amount)) + Decimal(str(r.loan_advance_deduction))
            net = Decimal(str(r.to_pay))

            gross_sum += gross
            deduction_sum += deduction
            net_sum += net

            att_summary = {
                "total_days": r.total_days,
                "full_days": r.full_day_count,
                "half_days": r.half_day_count,
                "off_days": r.off_day_count,
                "paid_leaves": float(r.paid_leave_count),
                "paid_days": float(r.paid_day_count),
                "unpaid_days": float(r.unpaid_day_count),
            }

            earnings_summary = {
                "daily_wage": float(r.daily_wage),
                "gross_wages": float(r.gross_wages),
                "overtime_amount": float(r.overtime_amount),
                "extras_amount": float(r.extras_amount),
                "gross_earnings": float(r.gross_earnings),
            }

            deduction_summary = {
                "penalties_amount": float(r.penalties_amount),
                "loan_deduction": float(r.loan_advance_deduction),
            }

            snapshot = {
                "employee_id": r.employee_id,
                "employee_code": emp_code,
                "employee_name": emp_name,
                "attendance": att_summary,
                "earnings": earnings_summary,
                "deductions": deduction_summary,
                "loan_amount": float(r.loan_advance_deduction),
                "arrears_amount": float(r.arrears_amount),
                "net_salary": float(r.to_pay),
            }

            employee_snapshots.append({
                "employee_id": r.employee_id,
                "attendance_summary": att_summary,
                "earnings_summary": earnings_summary,
                "deduction_summary": deduction_summary,
                "loan_amount": Decimal(str(r.loan_advance_deduction)),
                "arrears_amount": Decimal(str(r.arrears_amount)),
                "net_salary": net,
                "json_snapshot": snapshot,
            })

        async with self.transaction():
            fin_obj = PayrollFinalization(
                org_id=org_id,
                payroll_group_id=payload.payroll_group_id,
                payroll_period_id=None,
                from_date=payload.from_date,
                to_date=payload.to_date,
                payroll_module=payload.payroll_module,
                employee_count=len(computed_items),
                gross_amount=gross_sum,
                deduction_amount=deduction_sum,
                net_payable=net_sum,
                finalized_amount=net_sum,
                paid_amount=None,
                paid_on=None,
                status="Finalized",
                finalized_by=user_id,
                finalized_on=utcnow(),
                remarks=payload.remarks,
            )
            self.session.add(fin_obj)
            await self.session.flush()

            for emp_snap in employee_snapshots:
                emp_obj = PayrollFinalizationEmployee(
                    payroll_finalization_id=fin_obj.id,
                    employee_id=emp_snap["employee_id"],
                    attendance_summary=emp_snap["attendance_summary"],
                    earnings_summary=emp_snap["earnings_summary"],
                    deduction_summary=emp_snap["deduction_summary"],
                    loan_amount=emp_snap["loan_amount"],
                    arrears_amount=emp_snap["arrears_amount"],
                    net_salary=emp_snap["net_salary"],
                    json_snapshot=emp_snap["json_snapshot"],
                )
                self.session.add(emp_obj)

            await self.session.flush()

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="finalization",
                action_type=ActionType.CREATE,
                title="Finalize Payroll Record",
                description=f"Finalized payroll for group '{group.name}' from {payload.from_date} to {payload.to_date} with {len(computed_items)} employees (Total: {net_sum}).",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        return await self.get_finalized_payroll_detail(org_id, fin_obj.id)

    async def list_finalized_payrolls(
        self,
        org_id: int,
        group_id: int | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
        status: str | None = None,
        page: int = 1,
        page_size: int = 25,
    ) -> PayrollFinalizationListResponse:
        """List paginated finalized payroll history records."""
        total = await self.finalizations.search_count(
            org_id=org_id,
            payroll_group_id=group_id,
            from_date=from_date,
            to_date=to_date,
            status=status,
        )
        records = await self.finalizations.search(
            org_id=org_id,
            payroll_group_id=group_id,
            from_date=from_date,
            to_date=to_date,
            status=status,
            page=page,
            page_size=page_size,
        )

        items = []
        for r in records:
            item = PayrollFinalizationResponseSchema.model_validate(r)
            item.payroll_group_name = r.payroll_group.name if r.payroll_group else None
            items.append(item)

        return self.paginate(items, page=page, page_size=page_size, total_records=total)

    async def get_finalized_payroll_detail(
        self,
        org_id: int,
        finalization_id: int,
    ) -> PayrollFinalizationResponseSchema:
        """Retrieve complete details for a finalized payroll including frozen employee snapshots."""
        fin = await self.finalizations.get_by_id_in_org(org_id, finalization_id)
        if not fin:
            raise NotFoundException(f"Finalized payroll record {finalization_id} not found.")

        resp = PayrollFinalizationResponseSchema.model_validate(fin)
        resp.payroll_group_name = fin.payroll_group.name if fin.payroll_group else None

        emp_ids = [e.employee_id for e in fin.employees]
        emp_map: dict[int, Employee] = {}
        if emp_ids:
            emp_rows = await self.employees.get_by_ids(org_id, emp_ids)
            emp_map = {e.employee_id: e for e in emp_rows}

        employees_list = []
        for e in fin.employees:
            emp_info = emp_map.get(e.employee_id)
            emp_code = emp_info.employee_code if emp_info else f"EMP{e.employee_id}"
            emp_name = emp_info.employee_name if emp_info else f"Employee {e.employee_id}"

            employees_list.append(
                PayrollFinalizationEmployeeSchema(
                    id=e.id,
                    payroll_finalization_id=e.payroll_finalization_id,
                    employee_id=e.employee_id,
                    employee_code=emp_code,
                    employee_name=emp_name,
                    attendance_summary=e.attendance_summary,
                    earnings_summary=e.earnings_summary,
                    deduction_summary=e.deduction_summary,
                    loan_amount=e.loan_amount,
                    arrears_amount=e.arrears_amount,
                    net_salary=e.net_salary,
                    json_snapshot=e.json_snapshot or {},
                    created_at=e.created_at,
                )
            )

        resp.employees = employees_list
        return resp

    async def pay_finalized_payroll(
        self,
        org_id: int,
        finalization_id: int,
        payload: PayrollFinalizationPaySchema,
        user_id: int,
    ) -> PayrollFinalizationResponseSchema:
        """Mark a finalized payroll as Paid and record disbursement information."""
        fin = await self.finalizations.get_by_id_in_org(org_id, finalization_id)
        if not fin:
            raise NotFoundException(f"Finalized payroll record {finalization_id} not found.")

        if fin.status == "Cancelled":
            raise PayrollValidationException("Cancelled payroll cannot be marked as paid.")

        pay_amt = payload.paid_amount if payload.paid_amount is not None else fin.finalized_amount
        pay_date = payload.paid_on if payload.paid_on is not None else utcnow()

        async with self.transaction():
            update_dict: dict[str, Any] = {
                "paid_amount": pay_amt,
                "paid_on": pay_date,
                "status": "Paid",
                "updated_at": utcnow(),
            }
            if payload.remarks:
                update_dict["remarks"] = f"{fin.remarks or ''}\n[Payment Note]: {payload.remarks}".strip()

            await self.finalizations.update(fin, update_dict)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="finalization",
                action_type=ActionType.UPDATE,
                title="Mark Payroll Paid",
                description=f"Marked finalized payroll ID {finalization_id} as Paid (Paid Amount: {pay_amt}).",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        return await self.get_finalized_payroll_detail(org_id, finalization_id)

    async def cancel_finalized_payroll(
        self,
        org_id: int,
        finalization_id: int,
        payload: PayrollFinalizationCancelSchema,
        user_id: int,
    ) -> PayrollFinalizationResponseSchema:
        """Cancel a finalized payroll record."""
        fin = await self.finalizations.get_by_id_in_org(org_id, finalization_id)
        if not fin:
            raise NotFoundException(f"Finalized payroll record {finalization_id} not found.")

        if fin.status == "Paid":
            raise PayrollValidationException("Paid payroll cannot be cancelled.")

        async with self.transaction():
            update_dict: dict[str, Any] = {
                "status": "Cancelled",
                "updated_at": utcnow(),
            }
            if payload.reason:
                update_dict["remarks"] = f"{fin.remarks or ''}\n[Cancelled]: {payload.reason}".strip()

            await self.finalizations.update(fin, update_dict)

            await self.audit.record(
                org_id=org_id,
                module="payroll",
                sub_module="finalization",
                action_type=ActionType.UPDATE,
                title="Cancel Finalized Payroll",
                description=f"Cancelled finalized payroll ID {finalization_id}. Reason: {payload.reason or 'N/A'}.",
                performed_by_user_id=user_id,
                performed_by_name=f"User {user_id}",
            )

        return await self.get_finalized_payroll_detail(org_id, finalization_id)

